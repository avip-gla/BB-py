"""BRESE data extraction and loading.

Extracts energy avoided, CO2 avoided, and cost-benefit data from the
BRESE-cost-benefit-analysis.xlsx workbook. Uses landmark-based row detection
to handle inconsistent layouts across state tabs.

Source: BRESE-cost-benefit-analysis.xlsx (11 state tabs)
Layout: Two groups with different header offsets; landmark search handles both.
"""
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from ecu.config import BRESE_DATA_DIR, BRESE_SOURCE_XLSX, STATE_TABS, STATE_INFO


def _find_row(ws, col: int, text: str, start: int = 1, end: int = 40) -> Optional[int]:
    """Find the row number where a cell in the given column contains the text."""
    for r in range(start, end + 1):
        val = ws.cell(r, col).value
        if val and isinstance(val, str) and text in val:
            return r
    return None


def _find_cell(ws, text: str, start_row: int = 1, end_row: int = 40,
               start_col: int = 1, end_col: int = 20) -> Optional[Tuple[int, int]]:
    """Find (row, col) of a cell containing the given text."""
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            val = ws.cell(r, c).value
            if val and isinstance(val, str) and text in val:
                return (r, c)
    return None


def extract_electricity_data(ws) -> pd.DataFrame:
    """Extract electricity avoided data from a state worksheet.

    Finds the "Total Electricity (MWh) Results" landmark, then reads
    year rows until a non-numeric year is encountered.

    Source: Excel state tabs, Electricity section (cols A-E).

    Returns:
        DataFrame with columns: year, annual_mwh_avoided, cumulative_mwh_avoided,
        annual_mt_co2_avoided, cumulative_mt_co2_avoided.
    """
    title_row = _find_row(ws, 1, "Total Electricity")
    if title_row is None:
        return pd.DataFrame()

    rows = []
    r = title_row + 1
    # Skip header row(s) and blank rows to find first data year
    while r <= ws.max_row:
        val = ws.cell(r, 1).value
        if isinstance(val, (int, float)) and 2020 <= val <= 2050:
            break
        r += 1

    # Read data rows
    while r <= ws.max_row:
        year_val = ws.cell(r, 1).value
        if not isinstance(year_val, (int, float)) or year_val < 2020 or year_val > 2050:
            break
        rows.append({
            "year": int(year_val),
            "annual_mwh_avoided": ws.cell(r, 2).value or 0,
            "cumulative_mwh_avoided": ws.cell(r, 3).value or 0,
            "annual_mt_co2_avoided": ws.cell(r, 4).value,
            "cumulative_mt_co2_avoided": ws.cell(r, 5).value,
        })
        r += 1

    df = pd.DataFrame(rows)

    # Fix: TN has CO2 data placed on the 2031 row instead of 2030.
    # All states should have CO2 at 2030. Move misplaced CO2 values.
    co2_rows = df[df["annual_mt_co2_avoided"].notna()]
    if len(co2_rows) == 1 and co2_rows.iloc[0]["year"] != 2030:
        src_idx = co2_rows.index[0]
        tgt_idx = df[df["year"] == 2030].index
        if len(tgt_idx) == 1:
            df.loc[tgt_idx[0], "annual_mt_co2_avoided"] = df.loc[src_idx, "annual_mt_co2_avoided"]
            df.loc[tgt_idx[0], "cumulative_mt_co2_avoided"] = df.loc[src_idx, "cumulative_mt_co2_avoided"]
            df.loc[src_idx, "annual_mt_co2_avoided"] = None
            df.loc[src_idx, "cumulative_mt_co2_avoided"] = None

    return df


def extract_ng_data(ws) -> pd.DataFrame:
    """Extract natural gas avoided data from a state worksheet.

    Finds the "Total Natural Gas (MMBtu) Results" landmark. Falls back to
    scanning for year values starting at row 25 (consistent across all tabs)
    if the label is missing (e.g., MS tab).

    Source: Excel state tabs, NG section (cols A-C).
    Note: CO2 columns (D, E) are never populated in the NG section.

    Returns:
        DataFrame with columns: year, annual_mmbtu_avoided, cumulative_mmbtu_avoided.
    """
    title_row = _find_row(ws, 1, "Total Natural Gas")

    if title_row is not None:
        r = title_row + 1
    else:
        # Fallback: NG data consistently starts at row 25
        r = 25

    # Skip to first data year
    while r <= ws.max_row:
        val = ws.cell(r, 1).value
        if isinstance(val, (int, float)) and 2020 <= val <= 2050:
            break
        r += 1

    rows = []
    while r <= ws.max_row:
        year_val = ws.cell(r, 1).value
        if not isinstance(year_val, (int, float)) or year_val < 2020 or year_val > 2050:
            break
        rows.append({
            "year": int(year_val),
            "annual_mmbtu_avoided": ws.cell(r, 2).value or 0,
            "cumulative_mmbtu_avoided": ws.cell(r, 3).value or 0,
        })
        r += 1

    return pd.DataFrame(rows)


def extract_cost_benefit(ws) -> dict:
    """Extract cost-benefit summary from a state worksheet.

    Uses landmark search for "Energy Cost Savings" to handle varying
    row/column positions across tabs.

    Source: Excel state tabs, Cost & Savings section.
    Values are NPV in millions $ through 2040.

    Returns:
        Dict with keys: savings_res, savings_com, savings_total,
        costs_res, costs_com, costs_total, bcr_res, bcr_com, bcr_total.
        Values are None if not found.
    """
    # Find savings row by landmark
    savings_pos = _find_cell(ws, "Energy Cost Savings")
    if savings_pos is None:
        return {}

    sr, sc = savings_pos
    # Residential, Commercial, Total are in the next 3 columns
    result = {
        "savings_res": ws.cell(sr, sc + 1).value,
        "savings_com": ws.cell(sr, sc + 2).value,
        "savings_total": ws.cell(sr, sc + 3).value,
    }

    # Costs row is the next row after savings
    costs_row = sr + 1
    result["costs_res"] = ws.cell(costs_row, sc + 1).value
    result["costs_com"] = ws.cell(costs_row, sc + 2).value
    result["costs_total"] = ws.cell(costs_row, sc + 3).value

    # Benefit-cost ratio is the row after costs
    bcr_row = costs_row + 1
    bcr_res = ws.cell(bcr_row, sc + 1).value
    bcr_com = ws.cell(bcr_row, sc + 2).value
    bcr_total = ws.cell(bcr_row, sc + 3).value

    # Only include BCR if it looks like a ratio (not a label)
    if isinstance(bcr_res, (int, float)):
        result["bcr_res"] = bcr_res
        result["bcr_com"] = bcr_com
        result["bcr_total"] = bcr_total

    return result


def extract_all_states(
    xlsx_path: Optional[str] = None,
) -> Dict[str, dict]:
    """Extract all data from the BRESE workbook for all state tabs.

    Args:
        xlsx_path: Path to the Excel file. Defaults to BRESE_SOURCE_XLSX.

    Returns:
        Dict mapping state tab name -> {
            "electricity": DataFrame,
            "ng": DataFrame,
            "cost_benefit": dict,
            "info": dict (from STATE_INFO),
        }
    """
    import openpyxl

    if xlsx_path is None:
        xlsx_path = BRESE_SOURCE_XLSX

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    results = {}

    for tab_name in STATE_TABS:
        ws = wb[tab_name]
        results[tab_name] = {
            "electricity": extract_electricity_data(ws),
            "ng": extract_ng_data(ws),
            "cost_benefit": extract_cost_benefit(ws),
            "info": STATE_INFO.get(tab_name, {}),
        }

    wb.close()
    return results


def export_to_csv(
    data: Optional[Dict[str, dict]] = None,
    output_dir: Optional[Path] = None,
) -> None:
    """Export extracted BRESE data to CSV files.

    Creates three CSV files:
      - electricity_avoided.csv: all states' electricity data
      - ng_avoided.csv: all states' natural gas data
      - cost_benefit_summary.csv: all states' cost-benefit data

    Args:
        data: Output from extract_all_states(). If None, extracts fresh.
        output_dir: Directory for CSVs. Defaults to BRESE_DATA_DIR.
    """
    if data is None:
        data = extract_all_states()
    if output_dir is None:
        output_dir = BRESE_DATA_DIR

    output_dir.mkdir(parents=True, exist_ok=True)

    # Electricity
    elec_frames = []
    for state, d in data.items():
        df = d["electricity"].copy()
        if not df.empty:
            df.insert(0, "state", state)
            elec_frames.append(df)
    if elec_frames:
        elec_all = pd.concat(elec_frames, ignore_index=True)
        elec_all.to_csv(output_dir / "electricity_avoided.csv", index=False)

    # Natural gas
    ng_frames = []
    for state, d in data.items():
        df = d["ng"].copy()
        if not df.empty:
            df.insert(0, "state", state)
            ng_frames.append(df)
    if ng_frames:
        ng_all = pd.concat(ng_frames, ignore_index=True)
        ng_all.to_csv(output_dir / "ng_avoided.csv", index=False)

    # Cost-benefit summary
    cb_rows = []
    for state, d in data.items():
        cb = d["cost_benefit"]
        if cb:
            row = {"state": state}
            row.update(cb)
            cb_rows.append(row)
    if cb_rows:
        cb_df = pd.DataFrame(cb_rows)
        cb_df.to_csv(output_dir / "cost_benefit_summary.csv", index=False)


def load_electricity_avoided(path: Optional[str] = None) -> pd.DataFrame:
    """Load electricity avoided data from CSV.

    Args:
        path: Optional path override.

    Returns:
        DataFrame with columns: state, year, annual_mwh_avoided,
        cumulative_mwh_avoided, annual_mt_co2_avoided, cumulative_mt_co2_avoided.
    """
    if path is None:
        path = BRESE_DATA_DIR / "electricity_avoided.csv"
    return pd.read_csv(path)


def load_ng_avoided(path: Optional[str] = None) -> pd.DataFrame:
    """Load natural gas avoided data from CSV.

    Args:
        path: Optional path override.

    Returns:
        DataFrame with columns: state, year, annual_mmbtu_avoided,
        cumulative_mmbtu_avoided.
    """
    if path is None:
        path = BRESE_DATA_DIR / "ng_avoided.csv"
    return pd.read_csv(path)


def load_cost_benefit(path: Optional[str] = None) -> pd.DataFrame:
    """Load cost-benefit summary from CSV.

    Args:
        path: Optional path override.

    Returns:
        DataFrame with columns: state, savings_res, savings_com, savings_total,
        costs_res, costs_com, costs_total, bcr_res, bcr_com, bcr_total.
    """
    if path is None:
        path = BRESE_DATA_DIR / "cost_benefit_summary.csv"
    return pd.read_csv(path)
