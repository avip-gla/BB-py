"""Add 'Transport (City-Specific)' tab to IAM_model.xlsx.

Uses the current Python pipeline: dynamic car/truck fractions from AEO LDV
sales shares, separate car/truck MPG, SPPC carbon intensity available directly,
last-match freight efficiency (average across weight classes), and freight_ehybrid
allocated to gasoline.

Usage:
    python scripts/build_transport_tabs.py
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from bau.config import (
    CITIES, CITY_REGION_MAP, CITY_STATE_MAP,
    PROJECTION_YEARS, CITY_AEO_SALES_REGION_MAP,
)
from bau.data_loader import load_all_data, get_carbon_intensity, get_ldv_sales_share
from bau.city import City
from bau.transport import (
    calculate_fuel_consumption, calculate_transport_emissions,
)


# ============================================================
# Shared styles
# ============================================================
title_font = Font(bold=True, size=14)
section_font = Font(bold=True, size=12, color="1F4E79")
note_font = Font(italic=True, size=10, color="666666")
bold_font = Font(bold=True, size=10)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, size=10, color="FFFFFF")
change_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
change_font = Font(bold=True, size=10, color="BF8F00")
num_fmt = "#,##0.0"
year_col_start = 3  # Column C


# ============================================================
# Shared helper functions
# ============================================================

def _write_year_headers(ws, row: int, proj_years: list) -> None:
    """Write year column headers with formatting."""
    for i, yr in enumerate(proj_years):
        c = ws.cell(row=row, column=year_col_start + i, value=yr)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")


def _write_data_row(ws, row: int, label: str, values: list,
                    fmt: str = "#,##0.0", lfont: Font = None) -> None:
    """Write a label + data values row."""
    c = ws.cell(row=row, column=1, value=label)
    if lfont:
        c.font = lfont
    for i, v in enumerate(values):
        if v is not None:
            c = ws.cell(row=row, column=year_col_start + i, value=v)
            c.number_format = fmt


def _write_documentation_header(ws, r: int, title: str, description: str,
                                changes: list, proj_years: list) -> int:
    """Write the documentation header section. Returns next row."""
    last_col = get_column_letter(year_col_start + len(proj_years) - 1)
    ws.merge_cells(f"A{r}:{last_col}{r}")
    ws[f"A{r}"] = title
    ws[f"A{r}"].font = title_font
    r += 1

    ws.merge_cells(f"A{r}:{last_col}{r}")
    ws[f"A{r}"] = description
    ws[f"A{r}"].font = note_font
    r += 2

    ws[f"A{r}"] = "CHANGES / NOTES"
    ws[f"A{r}"].font = section_font
    r += 1

    for label, desc in changes:
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = change_font
        ws[f"A{r}"].fill = change_fill
        ws[f"B{r}"] = desc
        ws[f"B{r}"].font = note_font
        r += 1

    r += 1
    return r


def _write_city_params_section(ws, r: int, city_results: dict,
                               ci_label: str = "CI Region Used") -> int:
    """Write the city input parameters table. Returns next row."""
    ws[f"A{r}"] = "CITY INPUT PARAMETERS"
    ws[f"A{r}"].font = section_font
    r += 1

    param_headers = ["City", "State", "Region", "FHWA Annual VMT (Base Year)", ci_label]
    for i, h in enumerate(param_headers):
        c = ws.cell(row=r, column=i + 1, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    r += 1

    for name in CITIES:
        cr = city_results[name]
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=cr["state"])
        ws.cell(row=r, column=3, value=cr["region"])
        c = ws.cell(row=r, column=4, value=cr["total_vmt_base"])
        c.number_format = "#,##0"
        ws.cell(row=r, column=5, value=cr["ci_region"])
        r += 1

    r += 1
    return r


def _write_total_emissions_section(ws, r: int, city_results: dict,
                                   proj_years: list) -> int:
    """Write total transport emissions by city section. Returns next row."""
    ws[f"A{r}"] = "TOTAL TRANSPORT EMISSIONS BY CITY (MT CO2)"
    ws[f"A{r}"].font = section_font
    r += 1

    ws.cell(row=r, column=1, value="City").font = header_font
    ws.cell(row=r, column=1).fill = header_fill
    ws.cell(row=r, column=2, value="State").font = header_font
    ws.cell(row=r, column=2).fill = header_fill
    _write_year_headers(ws, r, proj_years)
    r += 1

    for name in CITIES:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=city_results[name]["state"])
        by_year = city_results[name]["by_year"]
        for i, yr in enumerate(proj_years):
            if yr in by_year:
                c = ws.cell(row=r, column=year_col_start + i,
                            value=by_year[yr]["emissions"]["total_mt_co2"])
                c.number_format = num_fmt
        r += 1

    r += 1
    return r


def _write_emissions_by_fuel_section(ws, r: int, city_results: dict,
                                     proj_years: list) -> int:
    """Write emissions by fuel type per city section. Returns next row."""
    ws[f"A{r}"] = "EMISSIONS BY FUEL TYPE PER CITY (MT CO2)"
    ws[f"A{r}"].font = section_font
    r += 1

    fuel_keys = [
        ("gasoline_mt_co2", "Gasoline"),
        ("diesel_mt_co2", "Diesel"),
        ("ethanol_mt_co2", "Ethanol"),
        ("electricity_mt_co2", "Electricity"),
    ]

    for name in CITIES:
        ws[f"A{r}"] = name
        ws[f"A{r}"].font = Font(bold=True, size=11, underline="single")
        r += 1
        ws.cell(row=r, column=1, value="Fuel Type").font = header_font
        ws.cell(row=r, column=1).fill = header_fill
        _write_year_headers(ws, r, proj_years)
        r += 1

        by_year = city_results[name]["by_year"]

        vals = [by_year[yr]["emissions"]["total_mt_co2"] if yr in by_year else None
                for yr in proj_years]
        _write_data_row(ws, r, "Total Emissions", vals, lfont=bold_font)
        r += 1

        for key, label in fuel_keys:
            vals = [by_year[yr]["emissions"][key] if yr in by_year else None
                    for yr in proj_years]
            _write_data_row(ws, r, f"   {label}", vals)
            r += 1

        r += 1

    return r


def _write_fuel_consumption_section(ws, r: int, city_results: dict,
                                    proj_years: list) -> int:
    """Write fuel consumption per city section. Returns next row."""
    ws[f"A{r}"] = "FUEL CONSUMPTION PER CITY"
    ws[f"A{r}"].font = section_font
    r += 1

    consumption_keys = [
        ("gasoline_gallons", "Gasoline (gallons)"),
        ("diesel_gallons", "Diesel (gallons)"),
        ("ethanol_gallons", "Ethanol (gallons)"),
        ("electricity_mwh", "Electricity (MWh)"),
    ]

    for name in CITIES:
        ws[f"A{r}"] = name
        ws[f"A{r}"].font = Font(bold=True, size=11, underline="single")
        r += 1
        ws.cell(row=r, column=1, value="Fuel Type").font = header_font
        ws.cell(row=r, column=1).fill = header_fill
        _write_year_headers(ws, r, proj_years)
        r += 1

        by_year = city_results[name]["by_year"]
        for key, label in consumption_keys:
            vals = [by_year[yr]["fuel"][key] if yr in by_year else None
                    for yr in proj_years]
            _write_data_row(ws, r, f"   {label}", vals)
            r += 1

        r += 1

    return r


def _write_vmt_by_fuel_section(ws, r: int, city_results: dict,
                               proj_years: list) -> int:
    """Write VMT by fuel type per city section. Returns next row."""
    ws[f"A{r}"] = "VMT BY FUEL TYPE PER CITY"
    ws[f"A{r}"].font = section_font
    r += 1

    vmt_keys = [
        ("conventional_gasoline", "Conventional Gasoline"),
        ("tdi_diesel", "TDI Diesel"),
        ("flex_fuel", "Flex-Fuel"),
        ("electric", "Electric"),
        ("plugin_hybrid", "Plug-in Hybrid"),
        ("electric_hybrid", "Electric Hybrid"),
    ]

    for name in CITIES:
        ws[f"A{r}"] = name
        ws[f"A{r}"].font = Font(bold=True, size=11, underline="single")
        r += 1
        ws.cell(row=r, column=1, value="Fuel Type").font = header_font
        ws.cell(row=r, column=1).fill = header_fill
        _write_year_headers(ws, r, proj_years)
        r += 1

        by_year = city_results[name]["by_year"]

        vals = [by_year[yr]["vmt_total"] if yr in by_year else None
                for yr in proj_years]
        _write_data_row(ws, r, "Total VMT", vals, fmt="#,##0", lfont=bold_font)
        r += 1

        for key, label in vmt_keys:
            vals = [by_year[yr]["vmt_by_fuel"][key] if yr in by_year else None
                    for yr in proj_years]
            _write_data_row(ws, r, f"   {label}", vals, fmt="#,##0")
            r += 1

        r += 1

    return r


def _set_column_widths(ws, proj_years: list) -> None:
    """Set standard column widths and freeze panes."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    for i in range(len(proj_years)):
        ws.column_dimensions[get_column_letter(year_col_start + i)].width = 14
    ws.freeze_panes = "C1"


# ============================================================
# Compute function — uses current Python pipeline
# ============================================================

def compute_all_cities(all_data: dict) -> dict:
    """Run the current transport pipeline for all 25 cities.

    Uses dynamic car/truck fractions from AEO LDV sales shares,
    separate car/truck MPG (vehicle_class="car" vs "truck"),
    SPPC carbon intensity directly available (no fallback),
    freight efficiency uses last match (average across weight classes),
    and freight_ehybrid allocated to gasoline.
    """
    city_results = {}
    for name in CITIES:
        city = City(name=name, all_data=all_data)
        by_year = {}
        vmt_df = city._get_projected_vmt()
        total_vmt = city._get_city_vmt()

        for yr in PROJECTION_YEARS:
            year_row = vmt_df[vmt_df["year"] == yr]
            if year_row.empty:
                continue
            year_row = year_row.iloc[0]
            vmt_by_fuel = {
                "conventional_gasoline": year_row["vmt_conventional_gasoline"],
                "tdi_diesel": year_row["vmt_tdi_diesel"],
                "flex_fuel": year_row["vmt_flex_fuel"],
                "electric": year_row["vmt_electric"],
                "plugin_hybrid": year_row["vmt_plugin_hybrid"],
                "electric_hybrid": year_row["vmt_electric_hybrid"],
            }

            sales_region = CITY_AEO_SALES_REGION_MAP.get(name, "South Atlantic")
            car_fraction = get_ldv_sales_share(
                sales_region, "Cars", yr, all_data["aeo_ldv_sales"]
            )
            truck_fraction = get_ldv_sales_share(
                sales_region, "Pick Up Trucks", yr, all_data["aeo_ldv_sales"]
            )

            fuel = calculate_fuel_consumption(
                vmt_by_fuel, yr, all_data["aeo_mpg"],
                car_fraction=car_fraction,
                truck_fraction=truck_fraction,
                aeo_freight=all_data["aeo_freight"],
            )

            ci = get_carbon_intensity(city.region, yr, all_data["aeo_ci"])
            emissions = calculate_transport_emissions(fuel, ci)

            by_year[yr] = {
                "vmt_by_fuel": vmt_by_fuel,
                "vmt_total": sum(vmt_by_fuel.values()),
                "fuel": fuel,
                "emissions": emissions,
                "ci": ci,
            }

        city_results[name] = {
            "total_vmt_base": total_vmt,
            "region": CITY_REGION_MAP[name],
            "state": CITY_STATE_MAP[name],
            "ci_region": CITY_REGION_MAP[name],
            "by_year": by_year,
        }
    return city_results


# ============================================================
# Tab builder
# ============================================================

def build_transport_tab(wb: openpyxl.Workbook, city_results: dict) -> None:
    """Add the 'Transport (City-Specific)' sheet to the workbook."""
    tab_name = "Transport (City-Specific)"
    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(tab_name)

    proj_years = PROJECTION_YEARS

    # ---- Documentation Header ----
    changes = [
        ("City-specific VMT",
         "Each city uses own FHWA VMT (Excel used Atlanta for all)."),
        ("State-specific fuel mix",
         "Each city uses own state's AFDC registration shares."),
        ("Region-specific carbon intensity",
         "City's own AEO electricity market region (SPPC available directly)."),
        ("Car/truck MPG split",
         "Car MPG from AEO R9, truck from AEO R24."),
        ("Dynamic car/truck fractions",
         "From AEO LDV sales shares (R103-R107) by region and year."),
        ("Freight efficiency",
         "Uses average across weight classes (AEO R155-R160, last match in CSV)."),
        ("freight_ehybrid allocation",
         "Allocated to gasoline (matching Excel R13)."),
        ("KNOWN DIVERGENCE",
         "Excel R21 formula bug — uses diesel VMT for car flex-fuel "
         "instead of flex VMT. Python uses correct flex VMT."),
    ]

    r = _write_documentation_header(
        ws, 1,
        title="Transportation — City-Specific Emissions",
        description=(
            "City-specific transport emissions using dynamic car/truck fractions from "
            "AEO LDV sales shares, separate car/truck MPG values, SPPC carbon intensity "
            "available directly, last-match freight efficiency (average across weight classes), "
            "and corrected freight_ehybrid allocation to gasoline."
        ),
        changes=changes,
        proj_years=proj_years,
    )

    # ---- City Input Parameters ----
    r = _write_city_params_section(ws, r, city_results, ci_label="CI Region Used")

    # ---- Total Emissions by City ----
    r = _write_total_emissions_section(ws, r, city_results, proj_years)

    # ---- Emissions by Fuel Type per City ----
    r = _write_emissions_by_fuel_section(ws, r, city_results, proj_years)

    # ---- Fuel Consumption per City ----
    r = _write_fuel_consumption_section(ws, r, city_results, proj_years)

    # ---- VMT by Fuel Type per City ----
    r = _write_vmt_by_fuel_section(ws, r, city_results, proj_years)

    # ---- Python Module Documentation ----
    r = _write_python_module_docs(ws, r, proj_years)

    # ---- Column widths ----
    _set_column_widths(ws, proj_years)


def _write_python_module_docs(ws, r: int, proj_years: list) -> int:
    """Write Python module documentation section. Returns next row."""
    last_col = get_column_letter(year_col_start + len(proj_years) - 1)

    ws[f"A{r}"] = "PYTHON MODULE DOCUMENTATION"
    ws[f"A{r}"].font = section_font
    r += 1

    # --- transport.py ---
    ws[f"A{r}"] = "transport.py"
    ws[f"A{r}"].font = Font(bold=True, size=11, underline="single")
    r += 1
    transport_docs = [
        ("calculate_initial_vmt_by_fuel(total_vmt, state, afdc_shares)",
         "Allocate total VMT across fuel types using AFDC state shares. "
         "Maps to Excel Transport R45-R50."),
        ("project_vmt(initial_vmt_by_fuel, years)",
         "Project VMT forward using AEO growth rates. "
         "Maps to Excel Transport R45-R50 columns E+. "
         "Formula: VMT(yr) = VMT(yr-1) * (1 + rate)."),
        ("calculate_fuel_consumption(vmt_by_fuel, year, aeo_mpg, car_fraction, truck_fraction, aeo_freight)",
         "Convert VMT to fuel consumption. Car/truck MPG split via vehicle_class column. "
         "Freight uses last-match efficiency. "
         "Maps to Excel Transport R13-R16, R19-R38."),
        ("calculate_transport_emissions(fuel_consumption, carbon_intensity)",
         "Convert fuel to CO2 using EPA factors + regional CI. "
         "Maps to Excel Transport R7-R10."),
    ]
    for func, desc in transport_docs:
        ws[f"A{r}"] = func
        ws[f"A{r}"].font = Font(bold=True, size=9)
        ws[f"B{r}"] = desc
        ws[f"B{r}"].font = note_font
        r += 1
    r += 1

    # --- city.py ---
    ws[f"A{r}"] = "city.py"
    ws[f"A{r}"].font = Font(bold=True, size=11, underline="single")
    r += 1
    city_docs = [
        ("City.__init__(name, fixed_data, city_data, all_data)",
         "Initialize with region/state from config. Lazy-loads data if not provided."),
        ("City.transport_emissions(year)",
         "Full pipeline: FHWA VMT -> AFDC fuel split -> AEO growth -> "
         "fuel consumption (car/truck MPG split) -> emissions. "
         "Caches VMT projections in _transport_vmt_cache."),
        ("City._get_city_vmt()",
         "Lookup from FHWA data, multiplied by 1000. Maps to Excel Transport R44."),
        ("City._get_projected_vmt()",
         "Cached VMT projections for all years. Uses calculate_initial_vmt_by_fuel + project_vmt."),
    ]
    for func, desc in city_docs:
        ws[f"A{r}"] = func
        ws[f"A{r}"].font = Font(bold=True, size=9)
        ws[f"B{r}"] = desc
        ws[f"B{r}"].font = note_font
        r += 1
    r += 1

    # --- config.py ---
    ws[f"A{r}"] = "config.py"
    ws[f"A{r}"].font = Font(bold=True, size=11, underline="single")
    r += 1
    config_docs = [
        ("LDV_SHARE = 0.9, HDV_SHARE = 0.1",
         "Light/heavy duty vehicle share of VMT (Transport R53-R54)."),
        ("KWH_PER_GALLON_GASOLINE = 33.7",
         "Energy content conversion for EV MPGe (Transport R57)."),
        ("EMISSION_FACTORS_KG_CO2",
         "EPA emission factors: motor_gasoline=8.78, diesel=10.21, ethanol_100=5.75 (Transport R52-R63)."),
        ("VMT_GROWTH_RATES",
         "AEO 2025 Table 41 annual growth rates by fuel type (Transport R70-R86)."),
        ("CITY_REGION_MAP",
         "25 cities -> AEO electricity market regions."),
        ("CITY_STATE_MAP",
         "25 cities -> states (for AFDC lookups)."),
        ("CITY_AEO_SALES_REGION_MAP",
         "25 cities -> AEO census divisions (South Atlantic / Middle Atlantic) for LDV sales shares."),
    ]
    for const, desc in config_docs:
        ws[f"A{r}"] = const
        ws[f"A{r}"].font = Font(bold=True, size=9)
        ws[f"B{r}"] = desc
        ws[f"B{r}"].font = note_font
        r += 1
    r += 1

    # --- data_loader.py ---
    ws[f"A{r}"] = "data_loader.py"
    ws[f"A{r}"].font = Font(bold=True, size=11, underline="single")
    r += 1
    loader_docs = [
        ("load_all_data()",
         "Loads all CSVs into a single dict: fixed, aeo_ci, aeo_mpg, aeo_freight, "
         "aeo_ldv_sales, fhwa_vmt, afdc_shares, buildings_emissions, etc."),
        ("get_carbon_intensity(region, year, ci_df)",
         "Lookup MT CO2/MWh for a region and year from AEO carbon intensity table."),
        ("get_ldv_sales_share(region, vehicle_type, year, sales_df)",
         "Lookup car/truck fraction from AEO LDV sales shares table. "
         "Returns 0-1 fraction for 'Cars' or 'Pick Up Trucks'."),
        ("load_aeo_mpg()",
         "AEO MPG projections with vehicle_class column ('car'/'truck') to disambiguate."),
        ("load_aeo_freight_efficiency()",
         "AEO freight efficiency by weight class and fuel type. "
         "Average values are last rows in CSV (R155-R160)."),
    ]
    for func, desc in loader_docs:
        ws[f"A{r}"] = func
        ws[f"A{r}"].font = Font(bold=True, size=9)
        ws[f"B{r}"] = desc
        ws[f"B{r}"].font = note_font
        r += 1
    r += 1

    return r


# ============================================================
# Main
# ============================================================

def main():
    print("Loading data...")
    all_data = load_all_data()

    print("Computing city-specific transport emissions...")
    results = compute_all_cities(all_data)

    print("Building Excel tab...")
    xlsx_path = Path("IAM_model.xlsx")
    wb = openpyxl.load_workbook(xlsx_path)

    # Remove old v2/v3 tabs if they exist
    for old_tab in ["Transport (v2 — City-Specific)", "Transport (v3 — MPG Split)"]:
        if old_tab in wb.sheetnames:
            del wb[old_tab]
            print(f"  Removed old tab: {old_tab}")

    build_transport_tab(wb, results)
    wb.save(xlsx_path)

    # Print verification values
    for name in ["Atlanta"]:
        for yr in [2027]:
            val = results[name]["by_year"][yr]["emissions"]["total_mt_co2"]
            print(f"\n{name} {yr}: {val:,.0f} MT CO2")

    print(f"\nDone. Added 'Transport (City-Specific)' tab to {xlsx_path}")


if __name__ == "__main__":
    main()
