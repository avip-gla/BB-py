"""Add a 'Transport (City-Specific)' tab to IAM_model.xlsx.

Mirrors the structure of the original Transport tab but with city-specific
calculations for all 25 cities. Includes documentation of all changes.

Usage:
    python scripts/build_transport_tab.py
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from iam.config import (
    CITIES, CITY_REGION_MAP, CITY_STATE_MAP,
    PROJECTION_YEARS, BASE_YEAR, CITY_AEO_SALES_REGION_MAP,
)
from iam.data_loader import load_all_data, get_carbon_intensity, get_ldv_sales_share
from iam.city import City
from iam.transport import (
    calculate_initial_vmt_by_fuel, project_vmt,
    calculate_fuel_consumption, calculate_transport_emissions,
)


def compute_all_cities(all_data: dict) -> dict:
    """Run the city-specific transport pipeline for all 25 cities."""
    city_results = {}
    for name in CITIES:
        city = City(name=name, all_data=all_data)
        by_year = {}
        vmt_df = city._get_projected_vmt()
        total_vmt = city._get_city_vmt()

        for yr in PROJECTION_YEARS:
            year_row = vmt_df[vmt_df["year"] == yr]
            if year_row.empty:
                continue
            year_row = year_row.iloc[0]
            vmt_by_fuel = {
                "conventional_gasoline": year_row["vmt_conventional_gasoline"],
                "tdi_diesel": year_row["vmt_tdi_diesel"],
                "flex_fuel": year_row["vmt_flex_fuel"],
                "electric": year_row["vmt_electric"],
                "plugin_hybrid": year_row["vmt_plugin_hybrid"],
                "electric_hybrid": year_row["vmt_electric_hybrid"],
            }
            sales_region = CITY_AEO_SALES_REGION_MAP.get(name, "South Atlantic")
            car_fraction = get_ldv_sales_share(
                sales_region, "Cars", yr, all_data["aeo_ldv_sales"]
            )
            truck_fraction = get_ldv_sales_share(
                sales_region, "Pick Up Trucks", yr, all_data["aeo_ldv_sales"]
            )
            fuel = calculate_fuel_consumption(
                vmt_by_fuel, yr, all_data["aeo_mpg"],
                car_fraction=car_fraction,
                truck_fraction=truck_fraction,
                aeo_freight=all_data["aeo_freight"],
            )
            ci = get_carbon_intensity(city.region, yr, all_data["aeo_ci"])
            emissions = calculate_transport_emissions(fuel, ci)
            by_year[yr] = {
                "vmt_by_fuel": vmt_by_fuel,
                "vmt_total": sum(vmt_by_fuel.values()),
                "fuel": fuel,
                "emissions": emissions,
                "ci": ci,
            }
        city_results[name] = {
            "total_vmt_base": total_vmt,
            "region": CITY_REGION_MAP[name],
            "state": CITY_STATE_MAP[name],
            "by_year": by_year,
        }
    return city_results


def build_tab(wb: openpyxl.Workbook, city_results: dict) -> None:
    """Add the Transport (City-Specific) sheet to the workbook."""
    tab_name = "Transport (City-Specific)"
    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(tab_name)

    # ---- Styles ----
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12, color="1F4E79")
    note_font = Font(italic=True, size=10, color="666666")
    bold_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    change_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    change_font = Font(bold=True, size=10, color="BF8F00")
    num_fmt = "#,##0.0"

    proj_years = PROJECTION_YEARS
    year_col_start = 3  # Column C

    def write_year_headers(row: int) -> None:
        for i, yr in enumerate(proj_years):
            c = ws.cell(row=row, column=year_col_start + i, value=yr)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")

    def write_data_row(row: int, label: str, values: list,
                       fmt: str = "#,##0.0", lfont: Font = None) -> None:
        c = ws.cell(row=row, column=1, value=label)
        if lfont:
            c.font = lfont
        for i, v in enumerate(values):
            if v is not None:
                c = ws.cell(row=row, column=year_col_start + i, value=v)
                c.number_format = fmt

    # ================================================================
    # DOCUMENTATION HEADER
    # ================================================================
    r = 1
    last_col = get_column_letter(year_col_start + len(proj_years) - 1)
    ws.merge_cells(f"A{r}:{last_col}{r}")
    ws[f"A{r}"] = "Transportation — City-Specific Emissions"
    ws[f"A{r}"].font = title_font
    r += 1

    ws.merge_cells(f"A{r}:{last_col}{r}")
    ws[f"A{r}"] = (
        "This tab replaces the original Transport tab which calculated emissions "
        "for a single reference city (Atlanta) and applied those values to all 25 cities."
    )
    ws[f"A{r}"].font = note_font
    r += 2

    ws[f"A{r}"] = "CHANGES FROM ORIGINAL TRANSPORT TAB"
    ws[f"A{r}"].font = section_font
    r += 1

    changes = [
        ("CHANGE 1 — City-specific VMT",
         "Old: Single reference city VMT (Atlanta = 5,598,764,246). "
         "New: Each city uses its own FHWA VMT scaled to city proper via population ratio."),
        ("CHANGE 2 — State-specific fuel mix",
         "Old: Georgia AFDC vehicle shares for all cities. "
         "New: Each city uses its own state's AFDC registration shares to allocate VMT by fuel type."),
        ("CHANGE 3 — Region-specific carbon intensity",
         "Old: SRSE (Atlanta's region) carbon intensity for electricity emissions. "
         "New: Each city uses its own AEO electricity market region CI. "
         "SPPC data now available directly from AEO."),
        ("CHANGE 4 — Car/truck MPG split",
         "Old: Same MPG for cars and trucks, hardcoded 0.42/0.58 fraction. "
         "New: Car MPG from AEO R9, truck MPG from AEO R24. Car/truck fraction "
         "from AEO LDV sales shares (R103-R107) by region and year."),
        ("UNCHANGED — VMT growth rates",
         "AEO 2025 Table 41 growth rates per fuel type remain the same for all cities (national rates)."),
        ("UNCHANGED — Emission factors",
         "EPA emission factors (kg CO2/unit) are unchanged."),
    ]
    for label, desc in changes:
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = change_font
        ws[f"A{r}"].fill = change_fill
        ws[f"B{r}"] = desc
        ws[f"B{r}"].font = note_font
        r += 1

    r += 1

    # ================================================================
    # SECTION 1: City Input Parameters
    # ================================================================
    ws[f"A{r}"] = "CITY INPUT PARAMETERS"
    ws[f"A{r}"].font = section_font
    r += 1

    param_headers = ["City", "State", "Region", "FHWA Annual VMT (Base Year)", "CI Region Used"]
    for i, h in enumerate(param_headers):
        c = ws.cell(row=r, column=i + 1, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    r += 1

    for name in CITIES:
        cr = city_results[name]
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=cr["state"])
        ws.cell(row=r, column=3, value=cr["region"])
        c = ws.cell(row=r, column=4, value=cr["total_vmt_base"])
        c.number_format = "#,##0"
        ws.cell(row=r, column=5, value=cr["region"])
        r += 1

    r += 1

    # ================================================================
    # SECTION 2: Total Emissions by City
    # ================================================================
    ws[f"A{r}"] = "TOTAL TRANSPORT EMISSIONS BY CITY (MT CO2)"
    ws[f"A{r}"].font = section_font
    r += 1

    ws.cell(row=r, column=1, value="City").font = header_font
    ws.cell(row=r, column=1).fill = header_fill
    ws.cell(row=r, column=2, value="State").font = header_font
    ws.cell(row=r, column=2).fill = header_fill
    write_year_headers(r)
    r += 1

    for name in CITIES:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=city_results[name]["state"])
        by_year = city_results[name]["by_year"]
        for i, yr in enumerate(proj_years):
            if yr in by_year:
                c = ws.cell(row=r, column=year_col_start + i,
                            value=by_year[yr]["emissions"]["total_mt_co2"])
                c.number_format = num_fmt
        r += 1

    r += 1

    # ================================================================
    # SECTION 3: Emissions by Fuel Type per City
    # ================================================================
    ws[f"A{r}"] = "EMISSIONS BY FUEL TYPE PER CITY (MT CO2)"
    ws[f"A{r}"].font = section_font
    r += 1

    fuel_keys = [
        ("gasoline_mt_co2", "Gasoline"),
        ("diesel_mt_co2", "Diesel"),
        ("ethanol_mt_co2", "Ethanol"),
        ("electricity_mt_co2", "Electricity"),
    ]

    for name in CITIES:
        ws[f"A{r}"] = name
        ws[f"A{r}"].font = Font(bold=True, size=11, underline="single")
        r += 1
        ws.cell(row=r, column=1, value="Fuel Type").font = header_font
        ws.cell(row=r, column=1).fill = header_fill
        write_year_headers(r)
        r += 1

        by_year = city_results[name]["by_year"]

        vals = [by_year[yr]["emissions"]["total_mt_co2"] if yr in by_year else None
                for yr in proj_years]
        write_data_row(r, "Total Emissions", vals, lfont=bold_font)
        r += 1

        for key, label in fuel_keys:
            vals = [by_year[yr]["emissions"][key] if yr in by_year else None
                    for yr in proj_years]
            write_data_row(r, f"   {label}", vals)
            r += 1

        r += 1

    # ================================================================
    # SECTION 4: Fuel Consumption per City
    # ================================================================
    ws[f"A{r}"] = "FUEL CONSUMPTION PER CITY"
    ws[f"A{r}"].font = section_font
    r += 1

    consumption_keys = [
        ("gasoline_gallons", "Gasoline (gallons)"),
        ("diesel_gallons", "Diesel (gallons)"),
        ("ethanol_gallons", "Ethanol (gallons)"),
        ("electricity_mwh", "Electricity (MWh)"),
    ]

    for name in CITIES:
        ws[f"A{r}"] = name
        ws[f"A{r}"].font = Font(bold=True, size=11, underline="single")
        r += 1
        ws.cell(row=r, column=1, value="Fuel Type").font = header_font
        ws.cell(row=r, column=1).fill = header_fill
        write_year_headers(r)
        r += 1

        by_year = city_results[name]["by_year"]
        for key, label in consumption_keys:
            vals = [by_year[yr]["fuel"][key] if yr in by_year else None
                    for yr in proj_years]
            write_data_row(r, f"   {label}", vals)
            r += 1

        r += 1

    # ================================================================
    # SECTION 5: VMT by Fuel Type per City
    # ================================================================
    ws[f"A{r}"] = "VMT BY FUEL TYPE PER CITY"
    ws[f"A{r}"].font = section_font
    r += 1

    vmt_keys = [
        ("conventional_gasoline", "Conventional Gasoline"),
        ("tdi_diesel", "TDI Diesel"),
        ("flex_fuel", "Flex-Fuel"),
        ("electric", "Electric"),
        ("plugin_hybrid", "Plug-in Hybrid"),
        ("electric_hybrid", "Electric Hybrid"),
    ]

    for name in CITIES:
        ws[f"A{r}"] = name
        ws[f"A{r}"].font = Font(bold=True, size=11, underline="single")
        r += 1
        ws.cell(row=r, column=1, value="Fuel Type").font = header_font
        ws.cell(row=r, column=1).fill = header_fill
        write_year_headers(r)
        r += 1

        by_year = city_results[name]["by_year"]

        vals = [by_year[yr]["vmt_total"] if yr in by_year else None
                for yr in proj_years]
        write_data_row(r, "Total VMT", vals, fmt="#,##0", lfont=bold_font)
        r += 1

        for key, label in vmt_keys:
            vals = [by_year[yr]["vmt_by_fuel"][key] if yr in by_year else None
                    for yr in proj_years]
            write_data_row(r, f"   {label}", vals, fmt="#,##0")
            r += 1

        r += 1

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    for i in range(len(proj_years)):
        ws.column_dimensions[get_column_letter(year_col_start + i)].width = 14

    ws.freeze_panes = "C1"


def main():
    print("Loading data...")
    all_data = load_all_data()

    print("Computing city-specific transport for all 25 cities...")
    city_results = compute_all_cities(all_data)

    print("Building Excel tab...")
    xlsx_path = Path("IAM_model.xlsx")
    wb = openpyxl.load_workbook(xlsx_path)
    build_tab(wb, city_results)
    wb.save(xlsx_path)

    print(f"Done. Added 'Transport (City-Specific)' tab to {xlsx_path}")
    print(f"  Sections: Documentation, City Parameters, Total Emissions, "
          f"Emissions by Fuel, Fuel Consumption, VMT by Fuel")
    print(f"  Cities: {len(CITIES)}, Years: {PROJECTION_YEARS[0]}-{PROJECTION_YEARS[-1]}")


if __name__ == "__main__":
    main()
